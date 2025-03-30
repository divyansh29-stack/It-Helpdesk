from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_mail import Mail, Message
from datetime import datetime, timedelta
import pandas as pd
import os
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import uuid
import google.generativeai as genai
import openpyxl

load_dotenv()

# Create data directory if it doesn't exist
if not os.path.exists('data'):
    os.makedirs('data')

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'your-secret-key')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///helpdesk.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = os.getenv('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.getenv('MAIL_PASSWORD')

# Configure Google Generative AI
GOOGLE_API_KEY = os.getenv('GEMINI_API_KEY')
try:
    genai.configure(api_key=GOOGLE_API_KEY)
    # List available models
    for m in genai.list_models():
        print(f"Available model: {m.name}")
    
    model = genai.GenerativeModel('models/gemini-1.5-pro')
    # Test the model with a simple prompt
    response = model.generate_content("Test connection")
    if response and response.text:
        print("Gemini API configured successfully")
    else:
        print("Warning: Empty response from Gemini API")
        model = None
except Exception as e:
    print(f"Error configuring Gemini API: {str(e)}")
    model = None

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
mail = Mail(app)

# Database Models
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(20), nullable=False)  # 'admin', 'technician', 'employee'
    department = db.Column(db.String(50))
    designation = db.Column(db.String(50))
    employee_code = db.Column(db.String(20), unique=True)
    complaints = db.relationship('Complaint', backref='user', foreign_keys='Complaint.user_id')
    assigned_complaints = db.relationship('Complaint', backref='technician', foreign_keys='Complaint.technician_id')
    comments = db.relationship('Comment', backref='user', lazy=True)

class Complaint(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    complaint_no = db.Column(db.String(20), unique=True, nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    technician_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    issue = db.Column(db.Text, nullable=False)
    status = db.Column(db.String(20), default='Open')  # Open, In Progress, Resolved, Escalated
    priority = db.Column(db.String(20), default='Medium')  # Low, Medium, High
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    resolved_at = db.Column(db.DateTime)
    comments = db.relationship('Comment', backref='complaint', lazy=True)
    # New fields for employee details from chatbot
    employee_name = db.Column(db.String(100))
    employee_designation = db.Column(db.String(100))
    employee_department = db.Column(db.String(100))
    troubleshooting_steps = db.Column(db.Text)
    resolution_attempted = db.Column(db.Boolean, default=False)

class Comment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    complaint_id = db.Column(db.Integer, db.ForeignKey('complaint.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    content = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

# Routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        
        if user and user.password == password:  # In production, use proper password hashing
            login_user(user)
            if user.role == 'admin':
                return redirect(url_for('admin_dashboard'))
            elif user.role == 'technician':
                return redirect(url_for('technician_dashboard'))
            else:
                return redirect(url_for('employee_dashboard'))
        flash('Invalid credentials')
    return render_template('login.html')

@app.route('/chat')
@login_required
def chat_page():
    return render_template('chat.html')

@app.route('/api/chat', methods=['POST'])
@login_required
def chat_api():
    try:
        # Get JSON data from request
        data = request.json
        if not data or 'message' not in data:
            return jsonify({
                'response': 'I apologize, but I couldn\'t understand your message. Please try again.',
                'requiresComplaint': False
            }), 400
        
        message = data['message'].strip().lower()
        
        # Initialize session if not exists
        if 'chat_step' not in session:
            session['chat_step'] = 0
        
        # Check if Gemini API is available for more complex steps
        api_available = model is not None
        
        step = session.get('chat_step', 0)
        print(f"Current chat step: {step}, Message: {message}")
        
        if step == 0:
            if message == "hi" or message == "hello":
                session['chat_step'] = 1
                return jsonify({
                    'response': "Hello! I'm your IT Support Assistant. What is your name?",
                    'requiresComplaint': False
                })
            else:
                return jsonify({
                    'response': "Please type 'Hi' or 'Hello' to start the conversation.",
                    'requiresComplaint': False
                })
        elif step == 1:
            session['name'] = message
            session['chat_step'] = 2
            return jsonify({
                'response': f"Nice to meet you, {message}! Please enter your designation.",
                'requiresComplaint': False
            })
        elif step == 2:
            session['designation'] = message
            session['chat_step'] = 3
            return jsonify({
                'response': "Thank you. Now, please enter your department.",
                'requiresComplaint': False
            })
        elif step == 3:
            session['department'] = message
            session['chat_step'] = 4
            return jsonify({
                'response': "Please describe your IT problem in detail. What issues are you experiencing?",
                'requiresComplaint': False
            })
        elif step == 4:
            session['problem'] = message
            session['chat_step'] = 5
            
            print(f"Step 4: Problem set to: {message}")  # Debug log
            
            # Get troubleshooting steps from API
            try:
                troubleshooting_steps = search_gemini_api(message)
                
                # Check if API returned valid troubleshooting steps
                if not api_available or "apologize" in troubleshooting_steps.lower():
                    print("API unavailable, using fallback troubleshooting")
                    # Provide generic troubleshooting steps instead of immediately creating a ticket
                    fallback_steps = get_fallback_troubleshooting_steps(message)
                    session['last_resolution'] = fallback_steps
                    print("Fallback troubleshooting steps provided, asking if resolved")
                    return jsonify({
                        'response': f"Here are some troubleshooting steps:\n\n{fallback_steps}\n\nDid this resolve your issue? (Yes/No)",
                        'requiresComplaint': False
                    })
                
                session['last_resolution'] = troubleshooting_steps
                print("Troubleshooting steps provided, asking if resolved")
                return jsonify({
                    'response': f"Here are some troubleshooting steps:\n\n{troubleshooting_steps}\n\nDid this resolve your issue? (Yes/No)",
                    'requiresComplaint': False
                })
            except Exception as e:
                print(f"Error getting troubleshooting: {str(e)}")
                # Provide generic troubleshooting steps
                fallback_steps = get_fallback_troubleshooting_steps(message)
                session['last_resolution'] = fallback_steps
                print("Exception handled, using fallback troubleshooting")
                return jsonify({
                    'response': f"Here are some troubleshooting steps:\n\n{fallback_steps}\n\nDid this resolve your issue? (Yes/No)",
                    'requiresComplaint': False
                })
        elif step == 5:
            print(f"Step 5: User response to troubleshooting: {message}")  # Debug log
            if message == "yes":
                # Save resolved issue
                save_to_excel([
                    session.get('name', 'Unknown'),
                    session.get('designation', 'Unknown'),
                    session.get('department', 'Unknown'),
                    session.get('problem', 'Unknown'),
                    session.get('last_resolution', 'Unknown')
                ])
                session.clear()
                print("Issue resolved, session cleared")  # Debug log
                return jsonify({
                    'response': "Great! I'm glad your issue has been resolved. Your details have been saved, and you can always come back if you need more assistance.",
                    'requiresComplaint': False
                })
            elif message == "no":
                print("First solution didn't work, checking if API available for alternative")  # Debug log
                # If API is not available, try a different fallback solution
                try:
                    if not api_available:
                        print("API not available, using secondary fallback solution")  # Debug log
                        # Provide a more specific fallback solution as the second attempt
                        current_problem = session.get('problem', '')
                        secondary_fallback = get_secondary_fallback_steps(current_problem)
                        session['last_resolution'] = secondary_fallback
                        session['chat_step'] = 7  # Last attempt
                        print("Secondary fallback solution provided, asking if it worked")  # Debug log
                        return jsonify({
                            'response': f"Let's try these alternative steps instead:\n\n{secondary_fallback}\n\nDid this resolve your issue? (Yes/No)",
                            'requiresComplaint': False
                        })
                    
                    # Try alternative solution with a different prompt
                    alt_query = f"Alternative solution for: {session.get('problem', '')}"
                    troubleshooting_steps = search_gemini_api(alt_query)
                    
                    # Check if the API returned a proper response
                    if "apologize" in troubleshooting_steps.lower():
                        print("API couldn't find alternative solution, using secondary fallback")  # Debug log
                        # Provide a more specific fallback solution as the second attempt
                        current_problem = session.get('problem', '')
                        secondary_fallback = get_secondary_fallback_steps(current_problem)
                        session['last_resolution'] = secondary_fallback
                        session['chat_step'] = 7  # Last attempt
                        return jsonify({
                            'response': f"Let's try these alternative steps instead:\n\n{secondary_fallback}\n\nDid this resolve your issue? (Yes/No)",
                            'requiresComplaint': False
                        })
                    
                    session['last_resolution'] = troubleshooting_steps
                    session['chat_step'] = 7  # Last attempt
                    print("Alternative solution provided, asking if it worked")  # Debug log
                    return jsonify({
                        'response': f"Let's try this alternative solution instead:\n\n{troubleshooting_steps}\n\nDid this resolve your issue? (Yes/No)",
                        'requiresComplaint': False
                    })
                except Exception as e:
                    print(f"Error getting alternative solution: {str(e)}")
                    print("Exception when getting alternative solution, using secondary fallback")  # Debug log
                    current_problem = session.get('problem', '')
                    secondary_fallback = get_secondary_fallback_steps(current_problem)
                    session['last_resolution'] = secondary_fallback
                    session['chat_step'] = 7  # Last attempt
                    return jsonify({
                        'response': f"Let's try these alternative steps instead:\n\n{secondary_fallback}\n\nDid this resolve your issue? (Yes/No)",
                        'requiresComplaint': False
                    })
            else:
                return jsonify({
                    'response': "Please answer with 'Yes' or 'No'.",
                    'requiresComplaint': False
                })
        elif step == 6:
            print(f"Step 6: User response to alternative solution: {message}")  # Debug log
            if message == "yes":
                # Save resolved issue
                save_to_excel([
                    session.get('name', 'Unknown'),
                    session.get('designation', 'Unknown'),
                    session.get('department', 'Unknown'),
                    session.get('problem', 'Unknown'),
                    session.get('last_resolution', 'Unknown')
                ])
                session.clear()
                print("Alternative solution worked, session cleared")  # Debug log
                return jsonify({
                    'response': "Great! I'm glad the alternative solution worked. Your details have been saved, and you can always come back if you need more assistance.",
                    'requiresComplaint': False
                })
            elif message == "no":
                # Second attempt failed, automatically create a ticket
                print("Both troubleshooting attempts failed, creating ticket automatically")  # Debug log
                return create_support_ticket()
            else:
                return jsonify({
                    'response': "Please answer with 'Yes' or 'No'.",
                    'requiresComplaint': False
                })
        elif step == 7:
            print(f"Step 7: User response to secondary fallback: {message}")  # Debug log
            if message == "yes":
                # Save resolved issue
                save_to_excel([
                    session.get('name', 'Unknown'),
                    session.get('designation', 'Unknown'),
                    session.get('department', 'Unknown'),
                    session.get('problem', 'Unknown'),
                    session.get('last_resolution', 'Unknown')
                ])
                session.clear()
                print("Secondary fallback solution worked, session cleared")  # Debug log
                return jsonify({
                    'response': "Great! I'm glad the secondary fallback solution worked. Your details have been saved, and you can always come back if you need more assistance.",
                    'requiresComplaint': False
                })
            elif message == "no":
                # Second attempt failed, automatically create a ticket
                print("Both troubleshooting attempts failed, creating ticket automatically")  # Debug log
                return create_support_ticket()
            else:
                return jsonify({
                    'response': "Please answer with 'Yes' or 'No'.",
                    'requiresComplaint': False
                })
    
    except Exception as e:
        print(f"Error in chat_api: {str(e)}")
        session.clear()  # Clear session on error
        return jsonify({
            'response': 'I apologize, but I encountered an error. Please try again later.',
            'requiresComplaint': False
        }), 500

# Helper function to create a support ticket
def create_support_ticket():
    try:
        print("In create_support_ticket function")  # Debug log
        # Get all available technicians
        technicians = User.query.filter_by(role='technician').all()
        if not technicians:
            print("No technicians available")  # Debug log
            return jsonify({
                'response': "I apologize, but no technicians are available at the moment. Please try again later.",
                'requiresComplaint': False
            }), 500
        
        # Check if we have problem description
        problem = session.get('problem', '')
        if not problem:
            print("No problem description found in session")  # Debug log
            return jsonify({
                'response': "I apologize, but I couldn't determine your issue. Please try again.",
                'requiresComplaint': False
            }), 400
            
        print(f"Creating ticket for problem: {problem}")  # Debug log
        
        # Get user details from session
        user_name = session.get('name', current_user.username)
        user_designation = session.get('designation', current_user.designation)
        user_department = session.get('department', current_user.department)
        
        print(f"User details - Name: {user_name}, Designation: {user_designation}, Department: {user_department}")
        
        # Assign to technician with least active complaints
        assigned_technician = min(technicians, key=lambda t: len([c for c in t.assigned_complaints if c.status != 'Resolved']))
        
        # Determine priority based on issue type
        priority = 'Medium'
        if any(tag in problem.upper() for tag in ['[MEETING]', '[WEBINAR]', '[SEMINAR]']):
            priority = 'High'  # Meeting-related issues are higher priority
        
        # Create a new complaint
        complaint = Complaint(
            complaint_no=str(uuid.uuid4())[:8].upper(),
            user_id=current_user.id,
            technician_id=assigned_technician.id,
            issue=problem,
            status='Open',
            priority=priority,
            created_at=datetime.utcnow(),
            employee_name=user_name,
            employee_designation=user_designation,
            employee_department=user_department,
            troubleshooting_steps=session.get('last_resolution', ''),
            resolution_attempted=True
        )
        db.session.add(complaint)
        db.session.commit()
        
        print(f"Created complaint with ID: {complaint.id}, No: {complaint.complaint_no}")  # Debug log
        
        # Update Excel sheet
        complaint_data = {
            'complaint_no': complaint.complaint_no,
            'employee_name': user_name,
            'department': user_department,
            'employee_code': current_user.employee_code,
            'issue_description': problem,
            'status': complaint.status,
            'created_at': complaint.created_at,
            'resolved_at': None,
            'technician_name': assigned_technician.username,
            'resolution_time': None,
            'comments': f"Created through chatbot - Automatically created after troubleshooting failed"
        }
        update_excel_sheet(complaint_data)
        
        # Store ticket number to display to user
        ticket_no = complaint.complaint_no
        technician_name = assigned_technician.username
        
        print(f"Returning success response with ticket: {ticket_no}")  # Debug log
        
        # Clear session
        session.clear()
        
        # Provide detailed response with ticket information
        response_message = (
            f"Since the troubleshooting steps didn't resolve your issue, I've automatically created a support ticket for you:\n\n"
            f"📝 Ticket Number: {ticket_no}\n"
            f"👨‍💻 Assigned Technician: {technician_name}\n"
            f"🔍 Status: Open\n\n"
            f"You can view this ticket and its progress in your dashboard. The technician will contact you soon to resolve your issue."
        )
        
        return jsonify({
            'response': response_message,
            'requiresComplaint': False
        })
    except Exception as e:
        error_msg = str(e)
        print(f"Error creating complaint: {error_msg}")
        import traceback
        print(traceback.format_exc())  # Print full traceback
        session.clear()
        return jsonify({
            'response': "I apologize, but I encountered an error while creating your support ticket. Please try again later.",
            'requiresComplaint': False
        }), 500

def search_gemini_api(query):
    try:
        if model is None:
            return "I apologize, but I couldn't connect to the AI service. Please try again later or contact IT support directly."
            
        prompt = f"""As an IT Support Assistant, provide detailed troubleshooting steps for the following issue:
        {query}
        
        Please provide the steps in a clear, numbered format:
        1. First step
        2. Second step
        3. Third step
        etc.
        
        Also include:
        Common causes:
        - Cause 1
        - Cause 2
        - Cause 3
        
        When to contact IT support:
        - Situation 1
        - Situation 2
        - Situation 3
        
        IMPORTANT: Do not use any asterisks (*) or other special formatting in your response."""
        
        response = model.generate_content(prompt)
        if response and hasattr(response, 'text') and response.text:
            # Ensure the response is formatted correctly for display
            formatted_response = "Here are some troubleshooting steps:\n\n" + response.text
            return formatted_response
        else:
            print("Empty response from Gemini API")
            return "I apologize, but I couldn't generate a response. Please try again later or contact IT support directly."
    except Exception as e:
        print(f"Error in Gemini API: {str(e)}")
        return "I apologize, but there was an error connecting to the AI service. Please try again later or contact IT support directly."

def save_to_excel(data):
    file_path = 'data/user_data.xlsx'
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.active.append(["Name", "Designation", "Department", "Problem", "Resolution"])
        os.makedirs(os.path.dirname(file_path), exist_ok=True)

    sheet = workbook.active
    sheet.append(data)
    workbook.save(file_path)

def update_excel_sheet(complaint_data):
    """Update the Excel sheet with new complaint data"""
    excel_path = 'data/complaints_log.xlsx'
    
    # Create new Excel file if it doesn't exist
    if not os.path.exists(excel_path):
        df = pd.DataFrame(columns=[
            'Complaint No', 'Employee Name', 'Department', 'Employee Code',
            'Issue Description', 'Status', 'Created At', 'Resolved At',
            'Technician Name', 'Resolution Time (Hours)', 'Comments'
        ])
        df.to_excel(excel_path, index=False)
    
    # Load existing workbook
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # Add new row
    new_row = [
        complaint_data['complaint_no'],
        complaint_data['employee_name'],
        complaint_data['department'],
        complaint_data['employee_code'],
        complaint_data['issue_description'],
        complaint_data['status'],
        complaint_data['created_at'].strftime('%Y-%m-%d %H:%M:%S'),
        complaint_data['resolved_at'].strftime('%Y-%m-%d %H:%M:%S') if complaint_data['resolved_at'] else '',
        complaint_data['technician_name'] or '',
        complaint_data['resolution_time'] if complaint_data['resolved_at'] else '',
        complaint_data['comments']
    ]
    
    ws.append(new_row)
    
    # Apply styling
    for cell in ws[ws.max_row]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if ws.max_row % 2 == 0:
            cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    wb.save(excel_path)

@app.route('/api/chat/save', methods=['POST'])
@login_required
def save_chat():
    try:
        data = request.json
        if not data or 'issue' not in data:
            return jsonify({'error': 'Invalid request data'}), 400
        
        # Get all available technicians
        technicians = User.query.filter_by(role='technician').all()
        if not technicians:
            return jsonify({'error': 'No technicians available'}), 500
        
        # Assign to technician with least active complaints
        assigned_technician = min(technicians, key=lambda t: len([c for c in t.assigned_complaints if c.status != 'Resolved']))
        
        # Create a new complaint
        complaint = Complaint(
            complaint_no=str(uuid.uuid4())[:8].upper(),
            user_id=current_user.id,
            technician_id=assigned_technician.id,
            issue=data['issue'],
            status='Open',
            priority='Medium',
            created_at=datetime.utcnow(),
            employee_name=current_user.username,
            employee_designation=current_user.designation,
            employee_department=current_user.department,
            troubleshooting_steps=data['troubleshooting_steps'],
            resolution_attempted=True
        )
        db.session.add(complaint)
        db.session.commit()
        
        # Update Excel sheet
        complaint_data = {
            'complaint_no': complaint.complaint_no,
            'employee_name': current_user.username,
            'department': current_user.department,
            'employee_code': current_user.employee_code,
            'issue_description': data['issue'],
            'status': complaint.status,
            'created_at': complaint.created_at,
            'resolved_at': None,
            'technician_name': assigned_technician.username,
            'resolution_time': None,
            'comments': ''
        }
        update_excel_sheet(complaint_data)
        
        return jsonify({
            'complaintCreated': True,
            'complaintNo': complaint.complaint_no,
            'assignedTechnician': assigned_technician.username
        })
    except Exception as e:
        print(f"Error saving chat: {str(e)}")
        return jsonify({'error': 'Failed to save complaint'}), 500

@app.route('/complaint/create', methods=['GET', 'POST'])
@login_required
def create_complaint():
    # Redirect all attempts to manually create complaints to the chatbot
    flash('All IT support tickets must be created through the IT Support Assistant chatbot.', 'info')
    return redirect(url_for('chat_page'))

@app.route('/complaint/<int:complaint_id>')
@login_required
def view_complaint(complaint_id):
    try:
        print(f"Attempting to fetch complaint with ID: {complaint_id}")  # Debug log
        complaint = Complaint.query.get_or_404(complaint_id)
        print(f"Found complaint: {complaint.complaint_no}")  # Debug log
        
        # Check if user has permission to view this complaint
        if current_user.role == 'employee' and complaint.user_id != current_user.id:
            print("Unauthorized access attempt by employee")  # Debug log
            return jsonify({'error': 'Unauthorized'}), 403
        elif current_user.role == 'technician' and complaint.technician_id != current_user.id:
            print("Unauthorized access attempt by technician")  # Debug log
            return jsonify({'error': 'Unauthorized'}), 403
        
        # Prepare the response data with fallback values
        response_data = {
            'complaint_no': complaint.complaint_no,
            'issue': complaint.issue or 'No issue description available',
            'status': complaint.status or 'Unknown',
            'priority': complaint.priority or 'Unknown',
            'created_at': complaint.created_at.isoformat() if complaint.created_at else None,
            'technician': {
                'username': complaint.technician.username
            } if complaint.technician else None,
            'employee_name': complaint.employee_name or complaint.user.username or 'Unknown',
            'employee_designation': complaint.employee_designation or complaint.user.designation or 'N/A',
            'employee_department': complaint.employee_department or complaint.user.department or 'N/A',
            'troubleshooting_steps': complaint.troubleshooting_steps or 'No troubleshooting steps available',
            'resolution_attempted': complaint.resolution_attempted or False,
            'comments': [{
                'content': comment.content,
                'created_at': comment.created_at.isoformat() if comment.created_at else None,
                'user': {
                    'username': comment.user.username
                }
            } for comment in complaint.comments]
        }
        
        print(f"Returning complaint data: {response_data}")  # Debug log
        return jsonify(response_data)
        
    except Exception as e:
        print(f"Error in view_complaint route: {str(e)}")  # Debug log
        import traceback
        print(f"Traceback: {traceback.format_exc()}")  # Print full traceback
        return jsonify({'error': str(e)}), 500

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    if current_user.role != 'admin':
        return redirect(url_for('index'))
    
    # Get all complaints
    complaints = Complaint.query.order_by(Complaint.created_at.desc()).all()
    
    # Get all technicians
    technicians = User.query.filter_by(role='technician').all()
    
    # Get department statistics from chatbot data
    departments = {}
    for complaint in complaints:
        dept = complaint.employee_department if complaint.employee_department else complaint.user.department
        if dept not in departments:
            departments[dept] = {
                'name': dept,
                'total_issues': 0,
                'common_problems': [],
                'trend': 'stable'
            }
        departments[dept]['total_issues'] += 1
        
        # Add issue to common problems if it's not already there
        if complaint.issue not in departments[dept]['common_problems']:
            departments[dept]['common_problems'].append(complaint.issue)
    
    # Get hardware issues from chatbot data
    hardware_issues = {}
    for complaint in complaints:
        # Extract hardware type from issue description
        issue_lower = complaint.issue.lower()
        if 'laptop' in issue_lower or 'computer' in issue_lower:
            hw_type = 'Laptop/Computer'
        elif 'printer' in issue_lower:
            hw_type = 'Printer'
        elif 'network' in issue_lower or 'internet' in issue_lower:
            hw_type = 'Network'
        elif 'phone' in issue_lower or 'mobile' in issue_lower:
            hw_type = 'Phone/Mobile'
        else:
            hw_type = 'Other'
        
        if hw_type not in hardware_issues:
            hardware_issues[hw_type] = {
                'type': hw_type,
                'total_issues': 0,
                'common_problems': [],
                'resolution_rate': 0
            }
        hardware_issues[hw_type]['total_issues'] += 1
        
        # Add issue to common problems if it's not already there
        if complaint.issue not in hardware_issues[hw_type]['common_problems']:
            hardware_issues[hw_type]['common_problems'].append(complaint.issue)
    
    # Calculate resolution rates
    for hw_type in hardware_issues:
        resolved = sum(1 for c in complaints if c.status == 'Resolved' and 
                      (hw_type == 'Laptop/Computer' and ('laptop' in c.issue.lower() or 'computer' in c.issue.lower()) or
                       hw_type == 'Printer' and 'printer' in c.issue.lower() or
                       hw_type == 'Network' and ('network' in c.issue.lower() or 'internet' in c.issue.lower()) or
                       hw_type == 'Phone/Mobile' and ('phone' in c.issue.lower() or 'mobile' in c.issue.lower()) or
                       hw_type == 'Other' and not any(x in c.issue.lower() for x in ['laptop', 'computer', 'printer', 'network', 'internet', 'phone', 'mobile'])))
        total = hardware_issues[hw_type]['total_issues']
        hardware_issues[hw_type]['resolution_rate'] = round((resolved / total * 100) if total > 0 else 0)
    
    # Generate failure predictions based on chatbot data
    predictions = []
    for complaint in complaints:
        if complaint.employee_name and complaint.employee_department:
            # Analyze complaint history to predict potential failures
            issue_lower = complaint.issue.lower()
            if 'laptop' in issue_lower or 'computer' in issue_lower:
                hw_type = 'Laptop/Computer'
            elif 'printer' in issue_lower:
                hw_type = 'Printer'
            elif 'network' in issue_lower or 'internet' in issue_lower:
                hw_type = 'Network'
            elif 'phone' in issue_lower or 'mobile' in issue_lower:
                hw_type = 'Phone/Mobile'
            else:
                hw_type = 'Other'
            
            # Count similar issues for this employee
            similar_issues = sum(1 for c in complaints 
                               if c.employee_name == complaint.employee_name 
                               and c.employee_department == complaint.employee_department
                               and c.issue.lower() == complaint.issue.lower())
            
            # Determine risk level based on similar issues
            risk_level = 'High' if similar_issues > 3 else 'Medium' if similar_issues > 1 else 'Low'
            risk_level_color = 'danger' if risk_level == 'High' else 'warning' if risk_level == 'Medium' else 'success'
            
            # Add prediction if not already added for this employee
            if not any(p['employee'] == complaint.employee_name for p in predictions):
                predictions.append({
                    'employee': complaint.employee_name,
                    'department': complaint.employee_department,
                    'hardware': hw_type,
                    'predicted_failure': 'Within 30 days' if risk_level == 'High' else 'Within 90 days' if risk_level == 'Medium' else 'No immediate risk',
                    'risk_level': risk_level,
                    'risk_level_color': risk_level_color
                })
    
    return render_template('admin_dashboard.html',
                         complaints=complaints,
                         technicians=technicians,
                         departments=list(departments.values()),
                         hardware_issues=list(hardware_issues.values()),
                         predictions=predictions)

@app.route('/complaint/<int:complaint_id>/delete', methods=['POST'])
@login_required
def delete_complaint(complaint_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        complaint = Complaint.query.get_or_404(complaint_id)
        # Delete all associated comments first
        Comment.query.filter_by(complaint_id=complaint_id).delete()
        # Then delete the complaint
        db.session.delete(complaint)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/admin/add_user', methods=['POST'])
@login_required
def add_user():
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        data = request.json
        user = User(
            username=data['username'],
            email=data['email'],
            password=data['password'],  # In production, use proper password hashing
            role=data['role'],
            department=data['department'],
            designation=data['designation'],
            employee_code=data['employee_code']
        )
        db.session.add(user)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/technician/dashboard')
@login_required
def technician_dashboard():
    if current_user.role != 'technician':
        return redirect(url_for('index'))
    
    # Get assigned complaints
    complaints = Complaint.query.filter_by(technician_id=current_user.id).order_by(Complaint.created_at.desc()).all()
    return render_template('technician_dashboard.html', complaints=complaints)

@app.route('/employee/dashboard')
@login_required
def employee_dashboard():
    if current_user.role != 'employee':
        return redirect(url_for('index'))
    user_complaints = Complaint.query.filter_by(user_id=current_user.id).order_by(Complaint.created_at.desc()).all()
    return render_template('employee_dashboard.html', complaints=user_complaints)

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        role = request.form.get('role')
        department = request.form.get('department')
        designation = request.form.get('designation')
        employee_code = request.form.get('employee_code')
        
        if User.query.filter_by(username=username).first():
            flash('Username already exists')
            return redirect(url_for('register'))
        
        if User.query.filter_by(email=email).first():
            flash('Email already exists')
            return redirect(url_for('register'))
        
        if User.query.filter_by(employee_code=employee_code).first():
            flash('Employee code already exists')
            return redirect(url_for('register'))
        
        user = User(
            username=username,
            email=email,
            password=password,  # In production, use proper password hashing
            role=role,
            department=department,
            designation=designation,
            employee_code=employee_code
        )
        
        db.session.add(user)
        db.session.commit()
        
        flash('Registration successful! Please login.')
        return redirect(url_for('login'))
    
    return render_template('register.html')

def create_default_users():
    # Create admin user
    admin = User.query.filter_by(username='admin').first()
    if not admin:
        admin = User(
            username='admin',
            email='admin@company.com',
            password='admin123',
            role='admin',
            department='IT',
            designation='System Administrator',
            employee_code='ADM001'
        )
        db.session.add(admin)
    
    # Create 5 technicians
    technicians = [
        {'username': 'tech1', 'email': 'tech1@company.com', 'password': 'tech123', 'code': 'TECH001'},
        {'username': 'tech2', 'email': 'tech2@company.com', 'password': 'tech123', 'code': 'TECH002'},
        {'username': 'tech3', 'email': 'tech3@company.com', 'password': 'tech123', 'code': 'TECH003'},
        {'username': 'tech4', 'email': 'tech4@company.com', 'password': 'tech123', 'code': 'TECH004'},
        {'username': 'tech5', 'email': 'tech5@company.com', 'password': 'tech123', 'code': 'TECH005'}
    ]
    
    for tech in technicians:
        technician = User.query.filter_by(username=tech['username']).first()
        if not technician:
            technician = User(
                username=tech['username'],
                email=tech['email'],
                password=tech['password'],
                role='technician',
                department='IT Support',
                designation='IT Technician',
                employee_code=tech['code']
            )
            db.session.add(technician)
    
    # Create employee user
    employee = User.query.filter_by(username='emp1').first()
    if not employee:
        employee = User(
            username='emp1',
            email='emp1@company.com',
            password='emp123',
            role='employee',
            department='HR',
            designation='HR Executive',
            employee_code='EMP001'
        )
        db.session.add(employee)
    
    db.session.commit()

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out successfully.')
    return redirect(url_for('index'))

@app.route('/complaint/<int:complaint_id>/update_status', methods=['POST'])
@login_required
def update_complaint_status(complaint_id):
    complaint = Complaint.query.get_or_404(complaint_id)
    
    # Check if user has permission to update this complaint
    if current_user.role == 'technician' and complaint.technician_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        data = request.json
        complaint.status = data['status']
        if data['status'] == 'Resolved':
            complaint.resolved_at = datetime.utcnow()
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/complaint/<int:complaint_id>/assign_technician', methods=['POST'])
@login_required
def assign_technician(complaint_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    complaint = Complaint.query.get_or_404(complaint_id)
    try:
        data = request.json
        print(f"Received data: {data}")  # Debug log
        if not data or 'technician_id' not in data:
            return jsonify({'error': 'Missing technician_id in request'}), 400
            
        complaint.technician_id = data['technician_id']
        db.session.commit()
        print(f"Successfully assigned technician {data['technician_id']} to complaint {complaint_id}")  # Debug log
        return jsonify({'success': True})
    except Exception as e:
        print(f"Error assigning technician: {str(e)}")  # Debug log
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/complaint/<int:complaint_id>/add_comment', methods=['POST'])
@login_required
def add_comment(complaint_id):
    if current_user.role != 'technician':
        return jsonify({'error': 'Unauthorized'}), 403
    
    complaint = Complaint.query.get_or_404(complaint_id)
    if complaint.technician_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        data = request.json
        comment = Comment(
            content=data['content'],
            user_id=current_user.id,
            complaint_id=complaint_id
        )
        db.session.add(comment)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/admin/dashboard/stats')
@login_required
def admin_dashboard_stats():
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    complaints = Complaint.query.all()
    stats = {
        'total': len(complaints),
        'open': len([c for c in complaints if c.status == 'Open']),
        'in_progress': len([c for c in complaints if c.status == 'In Progress']),
        'resolved': len([c for c in complaints if c.status == 'Resolved'])
    }
    return jsonify(stats)

@app.route('/admin/export/complaints/excel')
@login_required
def export_complaints_excel():
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        complaints = Complaint.query.all()
        data = []
        for complaint in complaints:
            data.append({
                'Complaint No': complaint.complaint_no,
                'Employee Name': complaint.employee_name if complaint.employee_name else complaint.user.username,
                'Department & Code': f"{complaint.employee_department if complaint.employee_department else complaint.user.department}({complaint.user.employee_code})",
                'Issue': complaint.issue,
                'Status': complaint.status,
                'Priority': complaint.priority,
                'Created At': complaint.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'Resolved At': complaint.resolved_at.strftime('%Y-%m-%d %H:%M:%S') if complaint.resolved_at else '',
                'Assigned Technician': complaint.technician.username if complaint.technician else '',
                'Comments': '\n'.join([f"{comment.user.username}: {comment.content}" for comment in complaint.comments])
            })
        
        df = pd.DataFrame(data)
        excel_path = 'data/complaints_export.xlsx'
        df.to_excel(excel_path, index=False)
        
        return send_file(excel_path, as_attachment=True, download_name='complaints.xlsx')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/admin/export/complaints/csv')
@login_required
def export_complaints_csv():
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        complaints = Complaint.query.all()
        data = []
        for complaint in complaints:
            data.append({
                'Complaint No': complaint.complaint_no,
                'Employee Name': complaint.employee_name if complaint.employee_name else complaint.user.username,
                'Department & Code': f"{complaint.employee_department if complaint.employee_department else complaint.user.department}({complaint.user.employee_code})",
                'Issue': complaint.issue,
                'Status': complaint.status,
                'Priority': complaint.priority,
                'Created At': complaint.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'Resolved At': complaint.resolved_at.strftime('%Y-%m-%d %H:%M:%S') if complaint.resolved_at else '',
                'Assigned Technician': complaint.technician.username if complaint.technician else '',
                'Comments': '\n'.join([f"{comment.user.username}: {comment.content}" for comment in complaint.comments])
            })
        
        df = pd.DataFrame(data)
        csv_path = 'data/complaints_export.csv'
        df.to_csv(csv_path, index=False)
        
        return send_file(csv_path, as_attachment=True, download_name='complaints.csv')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/admin/export/complaints/pdf')
@login_required
def export_complaints_pdf():
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        complaints = Complaint.query.all()
        data = []
        for complaint in complaints:
            data.append({
                'Complaint No': complaint.complaint_no,
                'Employee Name': complaint.employee_name if complaint.employee_name else complaint.user.username,
                'Department & Code': f"{complaint.employee_department if complaint.employee_department else complaint.user.department}({complaint.user.employee_code})",
                'Issue': complaint.issue,
                'Status': complaint.status,
                'Priority': complaint.priority,
                'Created At': complaint.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'Resolved At': complaint.resolved_at.strftime('%Y-%m-%d %H:%M:%S') if complaint.resolved_at else '',
                'Assigned Technician': complaint.technician.username if complaint.technician else '',
                'Comments': '\n'.join([f"{comment.user.username}: {comment.content}" for comment in complaint.comments])
            })
        
        df = pd.DataFrame(data)
        pdf_path = 'data/complaints_export.pdf'
        
        # Create PDF using reportlab
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import Paragraph, Spacer
        
        doc = SimpleDocTemplate(pdf_path, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        # Add title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30
        )
        elements.append(Paragraph("Complaints Report", title_style))
        
        # Convert DataFrame to list of lists for the table
        table_data = [df.columns.tolist()] + df.values.tolist()
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        return send_file(pdf_path, as_attachment=True, download_name='complaints.pdf')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/complaint/<int:complaint_id>/update_priority', methods=['POST'])
@login_required
def update_complaint_priority(complaint_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    complaint = Complaint.query.get_or_404(complaint_id)
    
    try:
        data = request.json
        if not data or 'priority' not in data:
            return jsonify({'error': 'Missing priority in request'}), 400
            
        old_priority = complaint.priority
        complaint.priority = data['priority']
        db.session.commit()
        
        # Add a comment to notify about priority change
        priority_comment = Comment(
            content=f"Priority changed from {old_priority} to {data['priority']} by admin ({current_user.username})",
            user_id=current_user.id,
            complaint_id=complaint_id
        )
        db.session.add(priority_comment)
        db.session.commit()
        
        print(f"Successfully updated priority to {data['priority']} for complaint {complaint_id}")
        return jsonify({'success': True})
    except Exception as e:
        print(f"Error updating priority: {str(e)}")
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

# Function to provide fallback troubleshooting steps when API is unavailable
def get_fallback_troubleshooting_steps(problem):
    problem_lower = problem.lower()
    
    # PC/Computer issues
    if any(term in problem_lower for term in ['pc', 'computer', 'desktop', 'laptop', 'hanging', 'freeze', 'slow', 'crash']):
        return """1. Restart your computer completely
2. Close unnecessary applications and browser tabs
3. Check for available disk space (at least 10% should be free)
4. Run a quick virus scan
5. Update Windows and device drivers
6. Clear temporary files using Disk Cleanup
7. Check Task Manager (Ctrl+Shift+Esc) for programs using high resources

Common causes:
- Too many applications running simultaneously
- Outdated drivers or operating system
- Insufficient disk space
- Hardware issues

When to contact IT support:
- If the problem persists after trying these steps
- If you notice unusual behavior that might indicate malware
- If your computer repeatedly crashes with error messages"""

    # Network/Internet issues
    elif any(term in problem_lower for term in ['network', 'internet', 'wifi', 'connection', 'connect', 'disconnected']):
        return """1. Check if other devices can connect to the network
2. Restart your router/modem (unplug for 30 seconds, then plug back in)
3. Make sure Wi-Fi is enabled on your device
4. Try connecting with an Ethernet cable if possible
5. Forget the network and reconnect with the password
6. Run Windows network troubleshooter

Common causes:
- Router/modem issues
- Wi-Fi signal interference
- Network configuration problems
- ISP service outage

When to contact IT support:
- If multiple devices cannot connect
- If the network is unusually slow across all devices
- If you get error messages when trying to connect"""

    # Email issues
    elif any(term in problem_lower for term in ['email', 'outlook', 'mail', 'gmail', 'message']):
        return """1. Check your internet connection
2. Restart your email application
3. Verify your account settings are correct
4. Clear your email application cache
5. Check if you can access email via web browser
6. Ensure you haven't exceeded storage quota

Common causes:
- Connection issues
- Account configuration problems
- Temporary server outages
- Full mailbox

When to contact IT support:
- If you receive specific error codes
- If you can't access your email after trying these steps
- If you suspect your account has been compromised"""

    # Printer issues
    elif any(term in problem_lower for term in ['print', 'printer', 'scanning', 'scanner']):
        return """1. Check if the printer is powered on and connected to the network
2. Verify that there is paper in the tray and no paper jams
3. Restart the printer completely
4. Remove and re-add the printer on your computer
5. Update printer drivers
6. Try printing a test page

Common causes:
- Connection issues
- Driver problems
- Hardware malfunctions
- Paper jams or low ink/toner

When to contact IT support:
- If the printer displays error codes
- If print quality is consistently poor
- If the printer is making unusual noises"""

    # Software/Application issues
    elif any(term in problem_lower for term in ['software', 'application', 'program', 'app', 'not working']):
        return """1. Close and restart the application
2. Restart your computer
3. Check for application updates
4. Uninstall and reinstall the application
5. Clear the application cache if possible
6. Check if the application is compatible with your OS version

Common causes:
- Software bugs or glitches
- Corrupted installation
- Compatibility issues
- Insufficient system resources

When to contact IT support:
- If you receive specific error messages
- If reinstallation doesn't solve the problem
- If the application is mission-critical for your work"""

    # Login/Access issues
    elif any(term in problem_lower for term in ['login', 'password', 'access', 'account', 'authentication']):
        return """1. Verify you're using the correct username and password
2. Check if Caps Lock is turned on
3. Clear your browser cache and cookies
4. Try accessing from a different browser
5. Reset your password if you have self-service options
6. Check if the service is down for maintenance

Common causes:
- Incorrect credentials
- Expired passwords
- Account lockouts due to multiple failed attempts
- Browser cache issues

When to contact IT support:
- If you're locked out of your account
- If you can't reset your password
- If you suspect unauthorized access"""

    # Generic fallback
    else:
        return """1. Restart the device or application having issues
2. Check for any error messages and note them down
3. Verify your internet connection is working
4. Look for recent changes that might have caused the issue
5. Search for solutions in the company knowledge base
6. Try basic troubleshooting specific to the application

Common causes:
- Temporary system glitches
- Configuration issues
- Resource limitations
- Software bugs

When to contact IT support:
- If the issue persists after basic troubleshooting
- If you receive specific error codes
- If the issue is affecting your productivity
- If multiple users are experiencing the same problem"""

# Function to provide secondary fallback steps when API is unavailable
def get_secondary_fallback_steps(problem):
    problem_lower = problem.lower()
    
    # PC/Computer issues
    if any(term in problem_lower for term in ['pc', 'computer', 'desktop', 'laptop', 'hanging', 'freeze', 'slow', 'crash']):
        return """1. Check for Windows updates and install if available
2. Run System File Checker (SFC) by typing 'sfc /scannow' in Command Prompt
3. Check for hardware issues using built-in diagnostics
4. Try starting in Safe Mode to determine if it's a software conflict
5. Check Event Viewer for specific error codes
6. Perform a memory diagnostic test
7. Disconnect external devices and test again

This should help identify whether it's a hardware or software issue."""

    # Network/Internet issues
    elif any(term in problem_lower for term in ['network', 'internet', 'wifi', 'connection', 'connect', 'disconnected']):
        return """1. Reset TCP/IP stack by running 'netsh winsock reset' in Command Prompt
2. Release and renew your IP address using 'ipconfig /release' and 'ipconfig /renew'
3. Flush DNS cache with 'ipconfig /flushdns'
4. Change DNS settings to public DNS (like Google's 8.8.8.8 and 8.8.4.4)
5. Check for network adapter driver updates
6. Disable VPN or proxy if using one
7. Reset all network devices in the correct order (modem first, then router)

These steps address more advanced network configuration issues."""

    # Email issues
    elif any(term in problem_lower for term in ['email', 'outlook', 'mail', 'gmail', 'message']):
        return """1. Run Outlook in Safe Mode (hold Ctrl while launching)
2. Create a new Outlook profile and test with that
3. Check if your mailbox needs to be repaired with the Inbox Repair Tool (scanpst.exe)
4. Disable add-ins that might be causing issues
5. Check your email account settings for maximum size limits
6. Check if your email client is in offline mode
7. Verify your anti-virus isn't blocking email connections

These solutions target Outlook-specific issues and account configuration problems."""

    # Printer issues
    elif any(term in problem_lower for term in ['print', 'printer', 'scanning', 'scanner']):
        return """1. Clear the print queue (stop and restart Print Spooler service)
2. Check printer IP address and make sure it hasn't changed
3. Set a static IP for the printer if possible
4. Update printer firmware (check manufacturer website)
5. Check if printer needs calibration
6. Try a different USB port or cable if directly connected
7. Print directly to the device using its web interface if available

These steps help with more complex printer connection and driver issues."""

    # Software/Application issues
    elif any(term in problem_lower for term in ['software', 'application', 'program', 'app', 'not working']):
        return """1. Launch the application with admin privileges
2. Check for conflicts with anti-virus or firewall settings
3. Run the application in compatibility mode
4. Create a new user profile and test there
5. Check application logs for specific errors
6. Verify all dependencies are installed (like .NET Framework or Visual C++ Redistributables)
7. Try repairing the installation through Control Panel > Programs and Features

These steps help identify permission issues and software conflicts."""

    # Login/Access issues
    elif any(term in problem_lower for term in ['login', 'password', 'access', 'account', 'authentication']):
        return """1. Check if your account is locked due to too many failed attempts
2. Ensure your device time and date are accurate (important for authentication)
3. Check if you need to connect to VPN first before accessing certain systems
4. Try accessing from a different device to determine if it's device-specific
5. Check if multi-factor authentication needs to be reconfigured
6. Verify that your account hasn't expired or been disabled
7. Make sure you're using the correct domain if applicable (corporate vs. local account)

These steps address more complex authentication and account access issues."""

    # Generic fallback
    else:
        return """1. Take screenshots of any error messages for support reference
2. Check if colleagues are experiencing similar issues
3. Note when the issue started and any changes made around that time
4. Try using an alternative software/method temporarily if available
5. Check system requirements for the software you're using
6. Review recent updates that might have affected system behavior
7. Create a detailed document of when the issue occurs and steps to reproduce

Having this documentation will help support staff diagnose and fix the issue more quickly."""

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        create_default_users()
    app.run(debug=True) 