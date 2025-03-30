from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_mail import Mail, Message
from datetime import datetime, timedelta
import pandas as pd
import os
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import uuid
import google.generativeai as genai
import openpyxl

load_dotenv()

# Create data directory if it doesn't exist
if not os.path.exists('data'):
    os.makedirs('data')

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'your-secret-key')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///helpdesk.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = os.getenv('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.getenv('MAIL_PASSWORD')

# Configure Google Generative AI
GOOGLE_API_KEY = os.getenv('GEMINI_API_KEY')
try:
    genai.configure(api_key=GOOGLE_API_KEY)
    # List available models
    for m in genai.list_models():
        print(f"Available model: {m.name}")
    
    model = genai.GenerativeModel('models/gemini-1.5-pro')
    # Test the model with a simple prompt
    response = model.generate_content("Test connection")
    if response and response.text:
        print("Gemini API configured successfully")
    else:
        print("Warning: Empty response from Gemini API")
        model = None
except Exception as e:
    print(f"Error configuring Gemini API: {str(e)}")
    model = None

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
mail = Mail(app)

# Database Models
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(20), nullable=False)  # 'admin', 'technician', 'employee'
    department = db.Column(db.String(50))
    designation = db.Column(db.String(50))
    employee_code = db.Column(db.String(20), unique=True)
    complaints = db.relationship('Complaint', backref='user', foreign_keys='Complaint.user_id')
    assigned_complaints = db.relationship('Complaint', backref='technician', foreign_keys='Complaint.technician_id')
    comments = db.relationship('Comment', backref='user', lazy=True)

class Complaint(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    complaint_no = db.Column(db.String(20), unique=True, nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    technician_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    issue = db.Column(db.Text, nullable=False)
    status = db.Column(db.String(20), default='Open')  # Open, In Progress, Resolved, Escalated
    priority = db.Column(db.String(20), default='Medium')  # Low, Medium, High
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    resolved_at = db.Column(db.DateTime)
    comments = db.relationship('Comment', backref='complaint', lazy=True)
    # New fields for employee details from chatbot
    employee_name = db.Column(db.String(100))
    employee_designation = db.Column(db.String(100))
    employee_department = db.Column(db.String(100))
    troubleshooting_steps = db.Column(db.Text)
    resolution_attempted = db.Column(db.Boolean, default=False)

class Comment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    complaint_id = db.Column(db.Integer, db.ForeignKey('complaint.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    content = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

# Routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        
        if user and user.password == password:  # In production, use proper password hashing
            login_user(user)
            if user.role == 'admin':
                return redirect(url_for('admin_dashboard'))
            elif user.role == 'technician':
                return redirect(url_for('technician_dashboard'))
            else:
                return redirect(url_for('employee_dashboard'))
        flash('Invalid credentials')
    return render_template('login.html')

@app.route('/chat')
@login_required
def chat_page():
    return render_template('chat.html')

@app.route('/api/chat', methods=['POST'])
@login_required
def chat_api():
    try:
        data = request.json
        if not data or 'message' not in data:
            return jsonify({
                'response': 'I apologize, but I couldn\'t understand your message. Please try again.',
                'requiresComplaint': False
            }), 400
        
        message = data['message'].strip().lower()
        step = session.get('chat_step', 0)
        
        # Initialize session if not exists
        if 'chat_step' not in session:
            session['chat_step'] = 0
        
        # Check if Gemini API is available for more complex steps
        api_available = model is not None
        
        if step == 0:
            if message == "hi":
                session['chat_step'] = 1
                return jsonify({
                    'response': "Hello! I'm your IT Support Assistant. What is your name?",
                    'requiresComplaint': False
                })
            else:
                return jsonify({
                    'response': "Please type 'Hi' to start the conversation.",
                    'requiresComplaint': False
                })
        elif step == 1:
            session['name'] = message
            session['chat_step'] = 2
            return jsonify({
                'response': "Please enter your designation.",
                'requiresComplaint': False
            })
        elif step == 2:
            session['designation'] = message
            session['chat_step'] = 3
            return jsonify({
                'response': "Please enter your department.",
                'requiresComplaint': False
            })
        elif step == 3:
            session['department'] = message
            session['chat_step'] = 4
            return jsonify({
                'response': "Please describe your problem.",
                'requiresComplaint': False
            })
        elif step == 4:
            session['problem'] = message
            session['chat_step'] = 5
            
            # Get troubleshooting steps from API
            troubleshooting_steps = search_gemini_api(message)
            
            # If API failed, inform the user and offer to create a ticket
            if not api_available or "apologize" in troubleshooting_steps.lower():
                session['chat_step'] = 6
                session['last_resolution'] = "API unavailable"
                return jsonify({
                    'response': "I apologize, but I couldn't generate troubleshooting steps at this moment. Would you like to create a support ticket to have a technician assist you? (Yes/No)",
                    'requiresComplaint': False
                })
            
            session['last_resolution'] = troubleshooting_steps
            return jsonify({
                'response': f"Here are some troubleshooting steps:\n{troubleshooting_steps}\n\nDid this resolve your issue? (Yes/No)",
                'requiresComplaint': False
            })
        elif step == 5:
            if message == "yes":
                # Save resolved issue
                save_to_excel([
                    session.get('name', 'Unknown'),
                    session.get('designation', 'Unknown'),
                    session.get('department', 'Unknown'),
                    session.get('problem', 'Unknown'),
                    session.get('last_resolution', 'Unknown')
                ])
                session.clear()
                return jsonify({
                    'response': "Great! Your issue has been resolved. Your details have been saved.",
                    'requiresComplaint': False
                })
            elif message == "no":
                # If API is not available, skip the second attempt
                if not api_available:
                    session['chat_step'] = 6
                    return jsonify({
                        'response': "Would you like to create a support ticket to have a technician assist you? (Yes/No)",
                        'requiresComplaint': False
                    })
                
                # Try alternative solution with a different prompt
                try:
                    # Use a slightly different prompt to get alternative solutions
                    alt_query = f"Alternative solution for: {session.get('problem', '')}"
                    troubleshooting_steps = search_gemini_api(alt_query)
                    
                    # Check if the API returned a proper response
                    if "apologize" in troubleshooting_steps.lower():
                        session['chat_step'] = 6
                        return jsonify({
                            'response': "I couldn't find alternative solutions. Would you like to create a support ticket to have a technician assist you? (Yes/No)",
                            'requiresComplaint': False
                        })
                    
                    session['last_resolution'] = troubleshooting_steps
                    session['chat_step'] = 6
                    return jsonify({
                        'response': f"Try this instead:\n{troubleshooting_steps}\n\nDid this resolve your issue? (Yes/No)",
                        'requiresComplaint': False
                    })
                except Exception as e:
                    print(f"Error getting alternative solution: {str(e)}")
                    session['chat_step'] = 6
                    return jsonify({
                        'response': "I couldn't find alternative solutions. Would you like to create a support ticket to have a technician assist you? (Yes/No)",
                        'requiresComplaint': False
                    })
            else:
                return jsonify({
                    'response': "Please answer with 'Yes' or 'No'.",
                    'requiresComplaint': False
                })
        elif step == 6:
            if message == "yes":
                # If this is after the second troubleshooting attempt
                if session.get('last_resolution') != "API unavailable":
                    # Save resolved issue
                    save_to_excel([
                        session.get('name', 'Unknown'),
                        session.get('designation', 'Unknown'),
                        session.get('department', 'Unknown'),
                        session.get('problem', 'Unknown'),
                        session.get('last_resolution', 'Unknown')
                    ])
                    session.clear()
                    return jsonify({
                        'response': "Great! Your issue has been resolved. Your details have been saved.",
                        'requiresComplaint': False
                    })
                else:
                    # Create ticket if user wants help
                    return create_support_ticket()
            elif message == "no":
                # Create support ticket
                return create_support_ticket()
            else:
                return jsonify({
                    'response': "Please answer with 'Yes' or 'No'.",
                    'requiresComplaint': False
                })
    
    except Exception as e:
        print(f"Error in chat_api: {str(e)}")
        session.clear()  # Clear session on error
        return jsonify({
            'response': 'I apologize, but I encountered an error. Please try again later.',
            'requiresComplaint': False
        }), 500

# Helper function to create a support ticket
def create_support_ticket():
    try:
        # Get all available technicians
        technicians = User.query.filter_by(role='technician').all()
        if not technicians:
            return jsonify({
                'response': "I apologize, but no technicians are available at the moment. Please try again later.",
                'requiresComplaint': False
            }), 500
        
        # Assign to technician with least active complaints
        assigned_technician = min(technicians, key=lambda t: len([c for c in t.assigned_complaints if c.status != 'Resolved']))
        
        # Create a new complaint
        complaint = Complaint(
            complaint_no=str(uuid.uuid4())[:8].upper(),
            user_id=current_user.id,
            technician_id=assigned_technician.id,
            issue=session.get('problem', ''),
            status='Open',
            priority='Medium',
            created_at=datetime.utcnow(),
            employee_name=session.get('name', 'Unknown'),
            employee_designation=session.get('designation', 'Unknown'),
            employee_department=session.get('department', 'Unknown'),
            troubleshooting_steps=session.get('last_resolution', ''),
            resolution_attempted=True
        )
        db.session.add(complaint)
        db.session.commit()
        
        # Update Excel sheet
        complaint_data = {
            'complaint_no': complaint.complaint_no,
            'employee_name': current_user.username,
            'department': current_user.department,
            'employee_code': current_user.employee_code,
            'issue_description': session.get('problem', ''),
            'status': complaint.status,
            'created_at': complaint.created_at,
            'resolved_at': None,
            'technician_name': assigned_technician.username,
            'resolution_time': None,
            'comments': ''
        }
        update_excel_sheet(complaint_data)
        
        # Clear session
        session.clear()
        
        return jsonify({
            'response': f"I've created a support ticket (Ticket No: {complaint.complaint_no}) for your issue. A technician ({assigned_technician.username}) will be assigned to help you. You can track your complaint status in your dashboard.",
            'requiresComplaint': False
        })
    except Exception as e:
        print(f"Error creating complaint: {str(e)}")
        session.clear()
        return jsonify({
            'response': "I apologize, but I encountered an error while creating your support ticket. Please try again later.",
            'requiresComplaint': False
        }), 500

def search_gemini_api(query):
    try:
        if model is None:
            return "I apologize, but I couldn't connect to the AI service. Please try again later or contact IT support directly."
            
        prompt = f"""As an IT Support Assistant, provide detailed troubleshooting steps for the following issue:
        {query}
        
        Please provide the steps in a clear, numbered format:
        1. First step
        2. Second step
        3. Third step
        etc.
        
        Also include:
        Common causes:
        - Cause 1
        - Cause 2
        - Cause 3
        
        When to contact IT support:
        - Situation 1
        - Situation 2
        - Situation 3
        
        IMPORTANT: Do not use any asterisks (*) or other special formatting in your response."""
        
        response = model.generate_content(prompt)
        if response and hasattr(response, 'text') and response.text:
            # Ensure the response is formatted correctly for display
            formatted_response = "Here are some troubleshooting steps:\n\n" + response.text
            return formatted_response
        else:
            print("Empty response from Gemini API")
            return "I apologize, but I couldn't generate a response. Please try again later or contact IT support directly."
    except Exception as e:
        print(f"Error in Gemini API: {str(e)}")
        return "I apologize, but there was an error connecting to the AI service. Please try again later or contact IT support directly."

def save_to_excel(data):
    file_path = 'data/user_data.xlsx'
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.active.append(["Name", "Designation", "Department", "Problem", "Resolution"])
        os.makedirs(os.path.dirname(file_path), exist_ok=True)

    sheet = workbook.active
    sheet.append(data)
    workbook.save(file_path)

def update_excel_sheet(complaint_data):
    """Update the Excel sheet with new complaint data"""
    excel_path = 'data/complaints_log.xlsx'
    
    # Create new Excel file if it doesn't exist
    if not os.path.exists(excel_path):
        df = pd.DataFrame(columns=[
            'Complaint No', 'Employee Name', 'Department', 'Employee Code',
            'Issue Description', 'Status', 'Created At', 'Resolved At',
            'Technician Name', 'Resolution Time (Hours)', 'Comments'
        ])
        df.to_excel(excel_path, index=False)
    
    # Load existing workbook
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # Add new row
    new_row = [
        complaint_data['complaint_no'],
        complaint_data['employee_name'],
        complaint_data['department'],
        complaint_data['employee_code'],
        complaint_data['issue_description'],
        complaint_data['status'],
        complaint_data['created_at'].strftime('%Y-%m-%d %H:%M:%S'),
        complaint_data['resolved_at'].strftime('%Y-%m-%d %H:%M:%S') if complaint_data['resolved_at'] else '',
        complaint_data['technician_name'] or '',
        complaint_data['resolution_time'] if complaint_data['resolved_at'] else '',
        complaint_data['comments']
    ]
    
    ws.append(new_row)
    
    # Apply styling
    for cell in ws[ws.max_row]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if ws.max_row % 2 == 0:
            cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    wb.save(excel_path)

@app.route('/api/chat/save', methods=['POST'])
@login_required
def save_chat():
    try:
        data = request.json
        if not data or 'issue' not in data:
            return jsonify({'error': 'Invalid request data'}), 400
        
        # Get all available technicians
        technicians = User.query.filter_by(role='technician').all()
        if not technicians:
            return jsonify({'error': 'No technicians available'}), 500
        
        # Assign to technician with least active complaints
        assigned_technician = min(technicians, key=lambda t: len([c for c in t.assigned_complaints if c.status != 'Resolved']))
        
        # Create a new complaint
        complaint = Complaint(
            complaint_no=str(uuid.uuid4())[:8].upper(),
            user_id=current_user.id,
            technician_id=assigned_technician.id,
            issue=data['issue'],
            status='Open',
            priority='Medium',
            created_at=datetime.utcnow(),
            employee_name=current_user.username,
            employee_designation=current_user.designation,
            employee_department=current_user.department,
            troubleshooting_steps=data['troubleshooting_steps'],
            resolution_attempted=True
        )
        db.session.add(complaint)
        db.session.commit()
        
        # Update Excel sheet
        complaint_data = {
            'complaint_no': complaint.complaint_no,
            'employee_name': current_user.username,
            'department': current_user.department,
            'employee_code': current_user.employee_code,
            'issue_description': data['issue'],
            'status': complaint.status,
            'created_at': complaint.created_at,
            'resolved_at': None,
            'technician_name': assigned_technician.username,
            'resolution_time': None,
            'comments': ''
        }
        update_excel_sheet(complaint_data)
        
        return jsonify({
            'complaintCreated': True,
            'complaintNo': complaint.complaint_no,
            'assignedTechnician': assigned_technician.username
        })
    except Exception as e:
        print(f"Error saving chat: {str(e)}")
        return jsonify({'error': 'Failed to save complaint'}), 500

@app.route('/complaint/create', methods=['GET', 'POST'])
@login_required
def create_complaint():
    if request.method == 'POST':
        issue = request.form.get('issue')
        priority = request.form.get('priority', 'Medium')
        
        # Get all available technicians
        technicians = User.query.filter_by(role='technician').all()
        if not technicians:
            flash('No technicians available')
            return redirect(url_for('create_complaint'))
        
        # Assign to technician with least active complaints
        assigned_technician = min(technicians, key=lambda t: len([c for c in t.assigned_complaints if c.status != 'Resolved']))
        
        # Create a new complaint
        complaint = Complaint(
            complaint_no=str(uuid.uuid4())[:8].upper(),
            user_id=current_user.id,
            technician_id=assigned_technician.id,  # Automatically assign technician
            issue=issue,
            status='Open',
            priority=priority,
            created_at=datetime.utcnow(),
            employee_name=current_user.username,
            employee_designation=current_user.designation,
            employee_department=current_user.department,
            troubleshooting_steps='',
            resolution_attempted=False
        )
        db.session.add(complaint)
        db.session.commit()
        
        # Update Excel sheet
        complaint_data = {
            'complaint_no': complaint.complaint_no,
            'employee_name': current_user.username,
            'department': current_user.department,
            'employee_code': current_user.employee_code,
            'issue_description': issue,
            'status': complaint.status,
            'created_at': complaint.created_at,
            'resolved_at': None,
            'technician_name': assigned_technician.username,
            'resolution_time': None,
            'comments': ''
        }
        update_excel_sheet(complaint_data)
        
        flash('Complaint created successfully!')
        return redirect(url_for('employee_dashboard'))
    
    return render_template('create_complaint.html')

@app.route('/complaint/<int:complaint_id>')
@login_required
def view_complaint(complaint_id):
    try:
        print(f"Attempting to fetch complaint with ID: {complaint_id}")  # Debug log
        complaint = Complaint.query.get_or_404(complaint_id)
        print(f"Found complaint: {complaint.complaint_no}")  # Debug log
        
        # Check if user has permission to view this complaint
        if current_user.role == 'employee' and complaint.user_id != current_user.id:
            print("Unauthorized access attempt by employee")  # Debug log
            return jsonify({'error': 'Unauthorized'}), 403
        elif current_user.role == 'technician' and complaint.technician_id != current_user.id:
            print("Unauthorized access attempt by technician")  # Debug log
            return jsonify({'error': 'Unauthorized'}), 403
        
        # Prepare the response data with fallback values
        response_data = {
            'complaint_no': complaint.complaint_no,
            'issue': complaint.issue or 'No issue description available',
            'status': complaint.status or 'Unknown',
            'priority': complaint.priority or 'Unknown',
            'created_at': complaint.created_at.isoformat() if complaint.created_at else None,
            'technician': {
                'username': complaint.technician.username
            } if complaint.technician else None,
            'employee_name': complaint.employee_name or complaint.user.username or 'Unknown',
            'employee_designation': complaint.employee_designation or complaint.user.designation or 'N/A',
            'employee_department': complaint.employee_department or complaint.user.department or 'N/A',
            'troubleshooting_steps': complaint.troubleshooting_steps or 'No troubleshooting steps available',
            'resolution_attempted': complaint.resolution_attempted or False,
            'comments': [{
                'content': comment.content,
                'created_at': comment.created_at.isoformat() if comment.created_at else None,
                'user': {
                    'username': comment.user.username
                }
            } for comment in complaint.comments]
        }
        
        print(f"Returning complaint data: {response_data}")  # Debug log
        return jsonify(response_data)
        
    except Exception as e:
        print(f"Error in view_complaint route: {str(e)}")  # Debug log
        import traceback
        print(f"Traceback: {traceback.format_exc()}")  # Print full traceback
        return jsonify({'error': str(e)}), 500

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    if current_user.role != 'admin':
        return redirect(url_for('index'))
    
    # Get all complaints
    complaints = Complaint.query.order_by(Complaint.created_at.desc()).all()
    
    # Get all technicians
    technicians = User.query.filter_by(role='technician').all()
    
    # Get department statistics from chatbot data
    departments = {}
    for complaint in complaints:
        dept = complaint.employee_department if complaint.employee_department else complaint.user.department
        if dept not in departments:
            departments[dept] = {
                'name': dept,
                'total_issues': 0,
                'common_problems': [],
                'trend': 'stable'
            }
        departments[dept]['total_issues'] += 1
        
        # Add issue to common problems if it's not already there
        if complaint.issue not in departments[dept]['common_problems']:
            departments[dept]['common_problems'].append(complaint.issue)
    
    # Get hardware issues from chatbot data
    hardware_issues = {}
    for complaint in complaints:
        # Extract hardware type from issue description
        issue_lower = complaint.issue.lower()
        if 'laptop' in issue_lower or 'computer' in issue_lower:
            hw_type = 'Laptop/Computer'
        elif 'printer' in issue_lower:
            hw_type = 'Printer'
        elif 'network' in issue_lower or 'internet' in issue_lower:
            hw_type = 'Network'
        elif 'phone' in issue_lower or 'mobile' in issue_lower:
            hw_type = 'Phone/Mobile'
        else:
            hw_type = 'Other'
        
        if hw_type not in hardware_issues:
            hardware_issues[hw_type] = {
                'type': hw_type,
                'total_issues': 0,
                'common_problems': [],
                'resolution_rate': 0
            }
        hardware_issues[hw_type]['total_issues'] += 1
        
        # Add issue to common problems if it's not already there
        if complaint.issue not in hardware_issues[hw_type]['common_problems']:
            hardware_issues[hw_type]['common_problems'].append(complaint.issue)
    
    # Calculate resolution rates
    for hw_type in hardware_issues:
        resolved = sum(1 for c in complaints if c.status == 'Resolved' and 
                      (hw_type == 'Laptop/Computer' and ('laptop' in c.issue.lower() or 'computer' in c.issue.lower()) or
                       hw_type == 'Printer' and 'printer' in c.issue.lower() or
                       hw_type == 'Network' and ('network' in c.issue.lower() or 'internet' in c.issue.lower()) or
                       hw_type == 'Phone/Mobile' and ('phone' in c.issue.lower() or 'mobile' in c.issue.lower()) or
                       hw_type == 'Other' and not any(x in c.issue.lower() for x in ['laptop', 'computer', 'printer', 'network', 'internet', 'phone', 'mobile'])))
        total = hardware_issues[hw_type]['total_issues']
        hardware_issues[hw_type]['resolution_rate'] = round((resolved / total * 100) if total > 0 else 0)
    
    # Generate failure predictions based on chatbot data
    predictions = []
    for complaint in complaints:
        if complaint.employee_name and complaint.employee_department:
            # Analyze complaint history to predict potential failures
            issue_lower = complaint.issue.lower()
            if 'laptop' in issue_lower or 'computer' in issue_lower:
                hw_type = 'Laptop/Computer'
            elif 'printer' in issue_lower:
                hw_type = 'Printer'
            elif 'network' in issue_lower or 'internet' in issue_lower:
                hw_type = 'Network'
            elif 'phone' in issue_lower or 'mobile' in issue_lower:
                hw_type = 'Phone/Mobile'
            else:
                hw_type = 'Other'
            
            # Count similar issues for this employee
            similar_issues = sum(1 for c in complaints 
                               if c.employee_name == complaint.employee_name 
                               and c.employee_department == complaint.employee_department
                               and c.issue.lower() == complaint.issue.lower())
            
            # Determine risk level based on similar issues
            risk_level = 'High' if similar_issues > 3 else 'Medium' if similar_issues > 1 else 'Low'
            risk_level_color = 'danger' if risk_level == 'High' else 'warning' if risk_level == 'Medium' else 'success'
            
            # Add prediction if not already added for this employee
            if not any(p['employee'] == complaint.employee_name for p in predictions):
                predictions.append({
                    'employee': complaint.employee_name,
                    'department': complaint.employee_department,
                    'hardware': hw_type,
                    'predicted_failure': 'Within 30 days' if risk_level == 'High' else 'Within 90 days' if risk_level == 'Medium' else 'No immediate risk',
                    'risk_level': risk_level,
                    'risk_level_color': risk_level_color
                })
    
    return render_template('admin_dashboard.html',
                         complaints=complaints,
                         technicians=technicians,
                         departments=list(departments.values()),
                         hardware_issues=list(hardware_issues.values()),
                         predictions=predictions)

@app.route('/complaint/<int:complaint_id>/delete', methods=['POST'])
@login_required
def delete_complaint(complaint_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        complaint = Complaint.query.get_or_404(complaint_id)
        # Delete all associated comments first
        Comment.query.filter_by(complaint_id=complaint_id).delete()
        # Then delete the complaint
        db.session.delete(complaint)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/admin/add_user', methods=['POST'])
@login_required
def add_user():
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        data = request.json
        user = User(
            username=data['username'],
            email=data['email'],
            password=data['password'],  # In production, use proper password hashing
            role=data['role'],
            department=data['department'],
            designation=data['designation'],
            employee_code=data['employee_code']
        )
        db.session.add(user)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/technician/dashboard')
@login_required
def technician_dashboard():
    if current_user.role != 'technician':
        return redirect(url_for('index'))
    
    # Get assigned complaints
    complaints = Complaint.query.filter_by(technician_id=current_user.id).order_by(Complaint.created_at.desc()).all()
    return render_template('technician_dashboard.html', complaints=complaints)

@app.route('/employee/dashboard')
@login_required
def employee_dashboard():
    if current_user.role != 'employee':
        return redirect(url_for('index'))
    user_complaints = Complaint.query.filter_by(user_id=current_user.id).order_by(Complaint.created_at.desc()).all()
    return render_template('employee_dashboard.html', complaints=user_complaints)

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        role = request.form.get('role')
        department = request.form.get('department')
        designation = request.form.get('designation')
        employee_code = request.form.get('employee_code')
        
        if User.query.filter_by(username=username).first():
            flash('Username already exists')
            return redirect(url_for('register'))
        
        if User.query.filter_by(email=email).first():
            flash('Email already exists')
            return redirect(url_for('register'))
        
        if User.query.filter_by(employee_code=employee_code).first():
            flash('Employee code already exists')
            return redirect(url_for('register'))
        
        user = User(
            username=username,
            email=email,
            password=password,  # In production, use proper password hashing
            role=role,
            department=department,
            designation=designation,
            employee_code=employee_code
        )
        
        db.session.add(user)
        db.session.commit()
        
        flash('Registration successful! Please login.')
        return redirect(url_for('login'))
    
    return render_template('register.html')

def create_default_users():
    # Create admin user
    admin = User.query.filter_by(username='admin').first()
    if not admin:
        admin = User(
            username='admin',
            email='admin@company.com',
            password='admin123',
            role='admin',
            department='IT',
            designation='System Administrator',
            employee_code='ADM001'
        )
        db.session.add(admin)
    
    # Create 5 technicians
    technicians = [
        {'username': 'tech1', 'email': 'tech1@company.com', 'password': 'tech123', 'code': 'TECH001'},
        {'username': 'tech2', 'email': 'tech2@company.com', 'password': 'tech123', 'code': 'TECH002'},
        {'username': 'tech3', 'email': 'tech3@company.com', 'password': 'tech123', 'code': 'TECH003'},
        {'username': 'tech4', 'email': 'tech4@company.com', 'password': 'tech123', 'code': 'TECH004'},
        {'username': 'tech5', 'email': 'tech5@company.com', 'password': 'tech123', 'code': 'TECH005'}
    ]
    
    for tech in technicians:
        technician = User.query.filter_by(username=tech['username']).first()
        if not technician:
            technician = User(
                username=tech['username'],
                email=tech['email'],
                password=tech['password'],
                role='technician',
                department='IT Support',
                designation='IT Technician',
                employee_code=tech['code']
            )
            db.session.add(technician)
    
    # Create employee user
    employee = User.query.filter_by(username='emp1').first()
    if not employee:
        employee = User(
            username='emp1',
            email='emp1@company.com',
            password='emp123',
            role='employee',
            department='HR',
            designation='HR Executive',
            employee_code='EMP001'
        )
        db.session.add(employee)
    
    db.session.commit()

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out successfully.')
    return redirect(url_for('index'))

@app.route('/complaint/<int:complaint_id>/update_status', methods=['POST'])
@login_required
def update_complaint_status(complaint_id):
    complaint = Complaint.query.get_or_404(complaint_id)
    
    # Check if user has permission to update this complaint
    if current_user.role == 'technician' and complaint.technician_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        data = request.json
        complaint.status = data['status']
        if data['status'] == 'Resolved':
            complaint.resolved_at = datetime.utcnow()
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/complaint/<int:complaint_id>/assign_technician', methods=['POST'])
@login_required
def assign_technician(complaint_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    complaint = Complaint.query.get_or_404(complaint_id)
    try:
        data = request.json
        print(f"Received data: {data}")  # Debug log
        if not data or 'technician_id' not in data:
            return jsonify({'error': 'Missing technician_id in request'}), 400
            
        complaint.technician_id = data['technician_id']
        db.session.commit()
        print(f"Successfully assigned technician {data['technician_id']} to complaint {complaint_id}")  # Debug log
        return jsonify({'success': True})
    except Exception as e:
        print(f"Error assigning technician: {str(e)}")  # Debug log
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/complaint/<int:complaint_id>/add_comment', methods=['POST'])
@login_required
def add_comment(complaint_id):
    if current_user.role != 'technician':
        return jsonify({'error': 'Unauthorized'}), 403
    
    complaint = Complaint.query.get_or_404(complaint_id)
    if complaint.technician_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        data = request.json
        comment = Comment(
            content=data['content'],
            user_id=current_user.id,
            complaint_id=complaint_id
        )
        db.session.add(comment)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/admin/dashboard/stats')
@login_required
def admin_dashboard_stats():
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    complaints = Complaint.query.all()
    stats = {
        'total': len(complaints),
        'open': len([c for c in complaints if c.status == 'Open']),
        'in_progress': len([c for c in complaints if c.status == 'In Progress']),
        'resolved': len([c for c in complaints if c.status == 'Resolved'])
    }
    return jsonify(stats)

@app.route('/admin/export/complaints/excel')
@login_required
def export_complaints_excel():
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        complaints = Complaint.query.all()
        data = []
        for complaint in complaints:
            data.append({
                'Complaint No': complaint.complaint_no,
                'Employee Name': complaint.employee_name if complaint.employee_name else complaint.user.username,
                'Department & Code': f"{complaint.employee_department if complaint.employee_department else complaint.user.department}({complaint.user.employee_code})",
                'Issue': complaint.issue,
                'Status': complaint.status,
                'Priority': complaint.priority,
                'Created At': complaint.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'Resolved At': complaint.resolved_at.strftime('%Y-%m-%d %H:%M:%S') if complaint.resolved_at else '',
                'Assigned Technician': complaint.technician.username if complaint.technician else '',
                'Comments': '\n'.join([f"{comment.user.username}: {comment.content}" for comment in complaint.comments])
            })
        
        df = pd.DataFrame(data)
        excel_path = 'data/complaints_export.xlsx'
        df.to_excel(excel_path, index=False)
        
        return send_file(excel_path, as_attachment=True, download_name='complaints.xlsx')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/admin/export/complaints/csv')
@login_required
def export_complaints_csv():
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        complaints = Complaint.query.all()
        data = []
        for complaint in complaints:
            data.append({
                'Complaint No': complaint.complaint_no,
                'Employee Name': complaint.employee_name if complaint.employee_name else complaint.user.username,
                'Department & Code': f"{complaint.employee_department if complaint.employee_department else complaint.user.department}({complaint.user.employee_code})",
                'Issue': complaint.issue,
                'Status': complaint.status,
                'Priority': complaint.priority,
                'Created At': complaint.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'Resolved At': complaint.resolved_at.strftime('%Y-%m-%d %H:%M:%S') if complaint.resolved_at else '',
                'Assigned Technician': complaint.technician.username if complaint.technician else '',
                'Comments': '\n'.join([f"{comment.user.username}: {comment.content}" for comment in complaint.comments])
            })
        
        df = pd.DataFrame(data)
        csv_path = 'data/complaints_export.csv'
        df.to_csv(csv_path, index=False)
        
        return send_file(csv_path, as_attachment=True, download_name='complaints.csv')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/admin/export/complaints/pdf')
@login_required
def export_complaints_pdf():
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        complaints = Complaint.query.all()
        data = []
        for complaint in complaints:
            data.append({
                'Complaint No': complaint.complaint_no,
                'Employee Name': complaint.employee_name if complaint.employee_name else complaint.user.username,
                'Department & Code': f"{complaint.employee_department if complaint.employee_department else complaint.user.department}({complaint.user.employee_code})",
                'Issue': complaint.issue,
                'Status': complaint.status,
                'Priority': complaint.priority,
                'Created At': complaint.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'Resolved At': complaint.resolved_at.strftime('%Y-%m-%d %H:%M:%S') if complaint.resolved_at else '',
                'Assigned Technician': complaint.technician.username if complaint.technician else '',
                'Comments': '\n'.join([f"{comment.user.username}: {comment.content}" for comment in complaint.comments])
            })
        
        df = pd.DataFrame(data)
        pdf_path = 'data/complaints_export.pdf'
        
        # Create PDF using reportlab
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import Paragraph, Spacer
        
        doc = SimpleDocTemplate(pdf_path, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        # Add title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30
        )
        elements.append(Paragraph("Complaints Report", title_style))
        
        # Convert DataFrame to list of lists for the table
        table_data = [df.columns.tolist()] + df.values.tolist()
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        return send_file(pdf_path, as_attachment=True, download_name='complaints.pdf')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        create_default_users()
    app.run(debug=True) 